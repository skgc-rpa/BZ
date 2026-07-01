import imaplib
import email
from email.header import decode_header
import re
import pandas as pd
import os
import shutil
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

# --- 엑셀 서식을 위한 openpyxl 모듈 추가 ---
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

# ----------------- 사용자 설정 영역 -----------------
EMAIL = os.environ.get("GMAIL_USER", "skgc.rpa@gmail.com")
APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "")  
SEARCH_KEYWORD = "BZ"

RECIPIENT_EMAILS = os.environ.get("RECIPIENT_EMAILS", "carly1206@sk.com")
CC_EMAILS = os.environ.get("CC_EMAILS", "jp_lee@sk.com")

TEMP_DIR = "./original_bz_files"      
CUMULATIVE_PATH = "./bz_cumulative_result.xlsx" 
ZIP_OUTPUT_PATH = "./bz_original_files"        
# ---------------------------------------------------

def send_result_email(sender, app_pwd, recipients_str, cc_str, attachments):
    print("\n[안내] 결과 파일 이메일 전송을 시작합니다...")
    
    msg = MIMEMultipart()
    msg['From'] = sender
    msg['To'] = recipients_str       
    if cc_str:
        msg['Cc'] = cc_str           
    msg['Subject'] = "[RPA] BZ 데이터 누적 취합 결과"
    
    body = "BZ 데이터 누적 취합 엑셀 파일 송부 드립니다.\n\n업무에 참고하시기 바랍니다."
    msg.attach(MIMEText(body, 'plain'))
    
    for file_path in attachments:
        if os.path.exists(file_path):
            filename = os.path.basename(file_path)
            with open(file_path, "rb") as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
            
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f"attachment; filename={filename}")
            msg.attach(part)
        else:
            print(f"[경고] 첨부할 파일을 찾을 수 없습니다: {file_path}")
            
    try:
        server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        server.login(sender, app_pwd)
        server.send_message(msg)
        server.quit()
        print(f"[성공] 이메일 발송 완료 (수신: {recipients_str} / 참조: {cc_str})")
    except Exception as e:
        print(f"[오류] 이메일 발송에 실패했습니다: {e}")

def aggregate_zip_and_send_bz_data():
    os.makedirs(TEMP_DIR, exist_ok=True)

    try:
        imap = imaplib.IMAP4_SSL("imap.gmail.com")
        imap.login(EMAIL, APP_PASSWORD)
        print("[성공] Gmail IMAP 서버에 로그인했습니다.")
    except Exception as e:
        print(f"[오류] 로그인 실패: {e}")
        return

    sent_box_names = ['"[Gmail]/Sent Mail"', '"[Gmail]/보낸 편지함"']
    is_connected = False
    for box in sent_box_names:
        if imap.select(box)[0] == 'OK':
            print(f"[성공] {box}에 연결되었습니다.")
            is_connected = True
            break
            
    if not is_connected:
        print("[오류] 보낸 편지함을 찾을 수 없습니다.")
        imap.logout()
        return

    status, search_data = imap.search(None, f'TEXT "{SEARCH_KEYWORD}"')
    mail_ids = search_data[0].split()
    print(f"[안내] 검색된 메일 수: {len(mail_ids)}개")

    data_frames = {}
    date_pattern = re.compile(r'bz_result_(\d{4}-\d{2}-\d{2})\.xlsx', re.IGNORECASE)

    for mail_id in mail_ids:
        res, msg_data = imap.fetch(mail_id, "(RFC822)")
        for response_part in msg_data:
            if isinstance(response_part, tuple):
                msg = email.message_from_bytes(response_part[1])
                
                for part in msg.walk():
                    if part.get_content_maintype() == 'multipart' or part.get('Content-Disposition') is None:
                        continue

                    filename = part.get_filename()
                    if filename:
                        filename_decoded, encoding = decode_header(filename)[0]
                        if isinstance(filename_decoded, bytes):
                            filename_decoded = filename_decoded.decode(encoding if encoding else "utf-8", errors="ignore")
                        
                        match = date_pattern.search(filename_decoded)
                        if match:
                            date_str = match.group(1)
                            file_data = part.get_payload(decode=True)
                            
                            if file_data:
                                filepath = os.path.join(TEMP_DIR, filename_decoded)
                                with open(filepath, "wb") as f:
                                    f.write(file_data)
                                
                                try:
                                    df = pd.read_excel(filepath, usecols=[0, 1])
                                    df.columns = ['Index', date_str]
                                    df.set_index('Index', inplace=True)
                                    data_frames[date_str] = df
                                    print(f"   [가상 환경 임시 저장 및 로드 완료] {filename_decoded}")
                                except Exception as e:
                                    print(f"   [오류] {filename_decoded} 처리 실패: {e}")

    imap.close()
    imap.logout()

    if data_frames:
        print("\n[안내] 데이터 병합을 시작합니다...")
        sorted_dates = sorted(data_frames.keys())
        sorted_dfs = [data_frames[d] for d in sorted_dates]
        
        final_df = pd.concat(sorted_dfs, axis=1)
        final_df.reset_index(inplace=True)
        
        # === 엑셀 서식 적용 및 저장 파트 ===
        with pd.ExcelWriter(CUMULATIVE_PATH, engine='openpyxl') as writer:
            final_df.to_excel(writer, index=False, sheet_name='Sheet1')
            worksheet = writer.sheets['Sheet1']
            
            # 서식 객체 정의
            gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            bold_font = Font(bold=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            
            max_col = final_df.shape[1]
            max_row = final_df.shape[0] + 1  # 헤더 포함 전체 행 수
            
            # 1. 열 너비 설정 (첫 열 28, 나머지 12)
            worksheet.column_dimensions['A'].width = 28
            for col_idx in range(2, max_col + 1):
                col_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[col_letter].width = 12
                
            # 2. 첫 행(헤더) 서식 지정: 회색 음영 및 볼드
            for cell in worksheet[1]:
                cell.fill = gray_fill
                cell.font = bold_font
                
            # 3. 데이터가 있는 전체 범위에 테두리 적용
            for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    cell.border = thin_border
                    
        print(f"[성공] 1. 서식이 적용된 취합 엑셀 파일 생성 완료: {CUMULATIVE_PATH}")
        # =================================
        
        print("[안내] 원본 파일 압축을 시작합니다...")
        shutil.make_archive(ZIP_OUTPUT_PATH, 'zip', TEMP_DIR)
        final_zip_path = f"{ZIP_OUTPUT_PATH}.zip"
        print(f"[성공] 2. 원본 파일 압축 완료: {final_zip_path}")
        
        attachments_to_send = [CUMULATIVE_PATH]
        # attachments_to_send = [CUMULATIVE_PATH, final_zip_path]
        send_result_email(EMAIL, APP_PASSWORD, RECIPIENT_EMAILS, CC_EMAILS, attachments_to_send)
        
        print("\n[완료] 모든 작업이 성공적으로 끝났습니다. GitHub Actions 세션이 종료되면 임시 파일은 자동 삭제됩니다.")
    else:
        print("\n[안내] 조건에 맞는 첨부파일을 찾지 못했습니다.")

if __name__ == "__main__":
    aggregate_zip_and_send_bz_data()
